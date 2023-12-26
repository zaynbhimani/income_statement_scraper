from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

import openpyxl as xl
from openpyxl import Workbook

from pathlib import Path
import time


def get_income_statement(ticker):
    url = (f"https://seekingalpha.com/symbol/{ticker}/income-statement")
    print(f"Entering {url}")

    driver = webdriver.Chrome()
    driver.get(url)

    try:
        while WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, "Press & Hold"))):
            print("Presence of human test...")
            human_verif_button = driver.find_element(By.LINK_TEXT, "Press & Hold")
            actions = ActionChains(driver)

            actions.click_and_hold(human_verif_button)
            actions.perform()
    except:
        print("No human verification block")

    try:
        section = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//section[@data-test-id='income-statement-card']")))
        table = WebDriverWait(section, 10).until(
            EC.presence_of_element_located((By.XPATH, "//table[@class='w-full shrink-0 border-separate border-spacing-0 print:table-fixed table-fixed']")))
    except:
        driver.quit()

    try:
        workbook = xl.load_workbook("Company_Income_Statements.xlsx")
    except:
        workbook = xl.Workbook()

    try:
        active_sheet = workbook[f"{ticker}"]
    except:
        active_sheet = workbook.create_sheet(title=f"{ticker}")

    years = table.find_elements(By.XPATH, "//tr/th[@scope='col']")
    for index, year in enumerate(years):
        if year.text:
            cell = active_sheet.cell(1, index + 1)
            cell.value = year.text[4:]
    workbook.save("Company_Income_Statements.xlsx")

    rows = table.find_elements(By.XPATH, "//tbody/tr[@style='height: 36px;']")
    for index, row in enumerate(rows):
        value_header = row.find_element(By.XPATH, "//th/div[@tabindex='0']").text
        head_cell = active_sheet.cell(index + 2, 1)
        head_cell.value = value_header

        columns = row.find_elements(By.XPATH, "//td[@colspan='0']")
        for j_index, column in enumerate(columns):
            dollar_value = column.text
            val_cell = active_sheet(index + 2, j_index + 2)
            val_cell.value = dollar_value
    workbook.save("Company_Income_Statements.xlsx")


if __name__ == "__main__":
    print("Please enter the stock ticker:")
    ticker = input(">").upper()

    get_income_statement(ticker)
